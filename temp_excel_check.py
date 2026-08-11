from openpyxl import load_workbook
path = 'Final_Database_Import.xlsx'
wb = load_workbook(path, read_only=True)
sheet = wb.active
rows = list(sheet.iter_rows(values_only=True))
print('sheet title:', sheet.title)
print('total rows:', len(rows))
if rows:
    headers = rows[0]
    print('headers:', headers)
    print('\nfirst 5 rows:')
    for i, row in enumerate(rows[1:6], start=1):
        print(i, row)
